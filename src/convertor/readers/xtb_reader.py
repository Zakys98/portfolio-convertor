from datetime import datetime
from enum import StrEnum
from pathlib import Path
from typing import Any

import openpyxl

from convertor.readers.reader import Reader
from convertor.currency import Currency
from convertor.report import XtbReport
from convertor.transaction import Transaction, TransactionType
from convertor.utils import date_to_string


class XtbAction(StrEnum):
    DEP = "deposit"
    DIV = "DIVIDENT"
    FFI = "Free-funds Interest"
    FFIT = "Free-funds Interest Tax"
    SD = "Stamp Duty"
    SP = "Stock purchase"
    SS = "Stock sale"
    WT = "Withholding Tax"
    # Unused
    CT = "close trade"
    TRA = "transfer"
    SECF = "Sec Fee"
    SW = "swap"


class XtbReader(Reader[XtbReport]):
    # Column indices for data extraction in OLD format
    START_COL = 2
    END_COL_OFFSET = -4
    CURRENCY_ROW_INDEX = 5
    WORKSHEET_OLD = "CASH OPERATION HISTORY"
    WORKSHEET_NEW = "Cash Operations"

    # Data field indices after slicing (OLD)
    TIME_INDEX = 0
    TICKER_INDEX = -2
    AMOUNT_INDEX = -1

    def _extract_currency(self, row: tuple[Any, ...]) -> Currency:
        """Attempts to find the currency code in the specific header row."""
        try:
            currency_val = row[self.START_COL : self.END_COL_OFFSET][-2]
            return Currency(str(currency_val))
        except (IndexError, ValueError):
            return Currency.EUR

    def read(self, input_file: Path) -> XtbReport:
        workbook = openpyxl.load_workbook(input_file)
        if self.WORKSHEET_NEW in workbook.sheetnames:
            return self._read_new_format(workbook, input_file)
        elif self.WORKSHEET_OLD in workbook.sheetnames:
            return self._read_old_format(workbook)
        else:
            return XtbReport()

    def _read_new_format(self, workbook: openpyxl.Workbook, input_file: Path) -> XtbReport:
        sheet = workbook[self.WORKSHEET_NEW]
        rows = list(sheet.iter_rows(values_only=True))
        
        # Try extract currency from filename
        currency_str = "EUR"
        if "_" in input_file.name:
            c = input_file.name.split("_")[0]
            if len(c) == 3:
                currency_str = c
        currency = Currency(currency_str)
        
        report = XtbReport()
        
        for row in rows[5:]: # Data starts at index 5
            if not row or not row[0]:
                continue
                
            action_raw = str(row[0]).strip().lower()
            ticker = str(row[1]) if row[1] else None
            time = row[3]
            amount = float(row[4]) if isinstance(row[4], (int, float)) else 0.0
            desc = str(row[6]) if len(row) > 6 and row[6] else None
            
            if not isinstance(time, datetime):
                continue
            date_str = date_to_string(time)
            
            if action_raw == "stock purchase" or action_raw == "stock sale":
                quantity = 0.0
                if desc:
                    quantity_str = (
                        desc.upper()
                        .removeprefix("OPEN BUY")
                        .removeprefix("CLOSE BUY")
                        .removeprefix("OPEN SELL")
                        .removeprefix("CLOSE SELL")
                        .replace("@", "")
                        .split()[0]
                    )
                    if "/" in quantity_str:
                        quantity_str = quantity_str.split("/")[0]
                    try:
                        quantity = float(quantity_str)
                    except ValueError:
                        pass

                total_price = abs(amount) if amount else 0.0
                share_price = round(total_price / quantity, 2) if quantity > 0 else 0.0

                tx_type = TransactionType.BUY if action_raw == "stock purchase" else TransactionType.SELL

                tx = Transaction(
                    date=date_str,
                    type=tx_type,
                    ticker=ticker,
                    quantity=quantity,
                    price=share_price,
                    currency=currency,
                    notes=desc
                )
                report.transactions.append(tx)

            elif action_raw == "dividend":
                tx = Transaction(
                    date=date_str,
                    type=TransactionType.CASH_IN,
                    ticker=ticker,
                    price=abs(amount),
                    currency=currency,
                    notes="Dividend"
                )
                report.transactions.append(tx)

            elif "deposit" in action_raw or "withdrawal" in action_raw:
                tx_type = TransactionType.CASH_IN if amount > 0 else TransactionType.CASH_OUT
                tx = Transaction(
                    date=date_str,
                    type=tx_type,
                    price=abs(amount),
                    currency=currency,
                    notes="Deposit/Withdrawal"
                )
                report.transactions.append(tx)

            elif action_raw in ("withholding tax", "free funds interest tax"):
                tx = Transaction(
                    date=date_str,
                    type=TransactionType.FEE,
                    ticker=ticker,
                    price=abs(amount),
                    currency=currency,
                    notes=action_raw.capitalize()
                )
                report.transactions.append(tx)

            elif action_raw in ("stamp duty", "sec fee"):
                tx = Transaction(
                    date=date_str,
                    type=TransactionType.FEE,
                    ticker=ticker,
                    price=abs(amount),
                    currency=currency,
                    notes=action_raw.capitalize()
                )
                report.transactions.append(tx)

            elif action_raw == "free funds interest":
                tx = Transaction(
                    date=date_str,
                    type=TransactionType.CASH_IN,
                    price=abs(amount),
                    currency=currency,
                    notes="Free-funds Interest"
                )
                report.transactions.append(tx)

        return report

    def _read_old_format(self, workbook: openpyxl.Workbook) -> XtbReport:
        sheet = workbook[self.WORKSHEET_OLD]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return XtbReport()

        currency = self._extract_currency(rows[self.CURRENCY_ROW_INDEX])
        report = XtbReport()

        for row in rows:
            data = row[self.START_COL : self.END_COL_OFFSET]

            if not data or not data[0] or data[0] == "Type":
                continue

            try:
                if not isinstance(data[0], str):
                    continue
                action = XtbAction(data[0])
            except ValueError:
                continue

            values = data[1:]

            if len(values) < 4:
                continue

            time = values[self.TIME_INDEX]
            ticker = values[self.TICKER_INDEX]
            amount = values[self.AMOUNT_INDEX]

            if not isinstance(time, datetime):
                continue

            date_str = date_to_string(time)
            ticker_str = str(ticker) if ticker else None
            amount_float = float(amount) if isinstance(amount, (int, float)) else 0.0

            if action == XtbAction.SP or action == XtbAction.SS:
                desc = values[1]
                quantity = 0.0
                if isinstance(desc, str):
                    quantity_str = (
                        desc.removeprefix("OPEN BUY")
                        .removeprefix("CLOSE BUY")
                        .replace("@", "")
                        .split()[0]
                    )
                    if "/" in quantity_str:
                        quantity_str = quantity_str.split("/")[0]
                    try:
                        quantity = float(quantity_str)
                    except ValueError:
                        pass

                total_price = abs(amount_float) if amount_float else 0.0
                share_price = round(total_price / quantity, 2) if quantity > 0 else 0.0

                tx_type = TransactionType.BUY if action == XtbAction.SP else TransactionType.SELL

                tx = Transaction(
                    date=date_str,
                    type=tx_type,
                    ticker=ticker_str,
                    quantity=quantity,
                    price=share_price,
                    currency=currency,
                    notes=str(desc) if desc else None
                )
                report.transactions.append(tx)

            elif action == XtbAction.DIV:
                tx = Transaction(
                    date=date_str,
                    type=TransactionType.CASH_IN,
                    ticker=ticker_str,
                    price=abs(amount_float),
                    currency=currency,
                    notes="Dividend"
                )
                report.transactions.append(tx)

            elif action == XtbAction.DEP:
                tx_type = TransactionType.CASH_IN if amount_float > 0 else TransactionType.CASH_OUT
                tx = Transaction(
                    date=date_str,
                    type=tx_type,
                    price=abs(amount_float),
                    currency=currency,
                    notes="Deposit/Withdrawal"
                )
                report.transactions.append(tx)

            elif action in (XtbAction.WT, XtbAction.FFIT):
                tx = Transaction(
                    date=date_str,
                    type=TransactionType.FEE,
                    ticker=ticker_str,
                    price=abs(amount_float),
                    currency=currency,
                    notes=str(action.value)
                )
                report.transactions.append(tx)

            elif action in (XtbAction.SD, XtbAction.SECF):
                tx = Transaction(
                    date=date_str,
                    type=TransactionType.FEE,
                    ticker=ticker_str,
                    price=abs(amount_float),
                    currency=currency,
                    notes=str(action.value)
                )
                report.transactions.append(tx)

            elif action == XtbAction.FFI:
                tx = Transaction(
                    date=date_str,
                    type=TransactionType.CASH_IN,
                    price=abs(amount_float),
                    currency=currency,
                    notes="Free-funds Interest"
                )
                report.transactions.append(tx)

        return report
