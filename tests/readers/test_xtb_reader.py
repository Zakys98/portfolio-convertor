import pytest
from datetime import datetime
from convertor.readers.xtb_reader import XtbReader
from convertor.currency import Currency


@pytest.fixture
def mock_xlsx(tmp_path):
    """Creates a temporary XLSX file for testing."""
    import openpyxl

    xlsx_file = tmp_path / "test_xtb.xlsx"
    workbook = openpyxl.Workbook()

    sheet = workbook.create_sheet("CASH OPERATION HISTORY")

    # Add header rows and currency info
    sheet.append([None, None, "Currency", None, None, None, "EUR"])
    sheet.append([None, None, None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, "USD", None, None, None, None, None])

    # Add column headers
    sheet.append([None, None, "Type", "Time", "Ticker", "Amount", "Price", "Total"])

    # Add test data
    sheet.append(
        [
            None,
            None,
            "Stock purchase",
            datetime(2023, 1, 2, 15, 0),
            "OPEN BUY 0.1465 @ 102.36",
            "AAPL",
            150.0,
            -1500.0,
            None,
            None,
            None,
            "This has to be there",
        ]
    )
    sheet.append(
        [
            None,
            None,
            "Stock sale",
            datetime(2023, 1, 3, 16, 0),
            "CLOSE BUY 0.0594/0.104 @ 333.80",
            "TSLA",
            200.0,
            1000.0,
        ]
    )
    sheet.append(
        [None, None, "DIVIDENT", datetime(2023, 1, 4, 9, 0), None, None, "MSFT", 5.50]
    )
    sheet.append(
        [None, None, "deposit", datetime(2023, 1, 1, 10, 0), None, None, None, 1000.1]
    )

    workbook.save(xlsx_file)
    return xlsx_file


def test_xtb_reader_parses_correctly(mock_xlsx):
    reader = XtbReader()

    report = reader.read(mock_xlsx)

    assert report.deposit == 1000.1
    assert len(report.buys) == 1
    assert report.buys[0].ticker == "AAPL"

    assert len(report.sells) == 1
    assert report.sells[0].ticker == "TSLA"

    assert len(report.dividends) == 1
    assert report.dividends[0].ticker == "MSFT"
    assert report.dividends[0].amount == 5.50
    assert report.dividends[0].currency == Currency.USD


def test_buy_transaction_dates(mock_xlsx):
    reader = XtbReader()
    report = reader.read(mock_xlsx)

    assert len(report.buys) == 1
    assert report.buys[0].time == "2023-01-02 15:00:00"
    assert report.buys[0].ticker == "AAPL"


def test_sell_transaction_dates(mock_xlsx):
    reader = XtbReader()
    report = reader.read(mock_xlsx)

    assert len(report.sells) == 1
    assert report.sells[0].time == "2023-01-03 16:00:00"
    assert report.sells[0].ticker == "TSLA"


def test_dividend_dates(mock_xlsx):
    reader = XtbReader()
    report = reader.read(mock_xlsx)

    assert len(report.dividends) == 1
    assert report.dividends[0].time == "2023-01-04 09:00:00"
    assert report.dividends[0].ticker == "MSFT"
    assert report.dividends[0].currency == Currency.USD


def test_currency_extraction(mock_xlsx):
    reader = XtbReader()

    report = reader.read(mock_xlsx)

    assert report.deposit_currency == Currency.USD


def test_empty_worksheet(tmp_path):
    import openpyxl

    xlsx_file = tmp_path / "empty_xtb.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("CASH OPERATION HISTORY")

    # Add minimal structure
    for i in range(10):
        sheet.append([None] * 8)

    workbook.save(xlsx_file)

    reader = XtbReader()
    report = reader.read(xlsx_file)

    assert report.deposit == 0
    assert len(report.buys) == 0
    assert len(report.sells) == 0
    assert len(report.dividends) == 0


def test_multiple_transactions_same_type(tmp_path):
    """Test handling multiple buy, sell, and dividend transactions."""
    import openpyxl

    xlsx_file = tmp_path / "multiple_transactions.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("CASH OPERATION HISTORY")

    # Add header rows (must have 12 columns to match data rows)
    sheet.append(
        [None, None, "Currency", None, None, None, "USD", None, None, None, None, None]
    )
    for _ in range(4):
        sheet.append([None] * 12)
    sheet.append(
        [None, None, None, None, None, None, "USD", None, None, None, None, None]
    )
    sheet.append(
        [
            None,
            None,
            "Type",
            "Time",
            "Ticker",
            "Amount",
            "Price",
            "Total",
            None,
            None,
            None,
            None,
        ]
    )

    # Add multiple buys
    sheet.append(
        [
            None,
            None,
            "Stock purchase",
            datetime(2023, 2, 1, 10, 30),
            "OPEN BUY 10.0 @ 150.00",
            "AAPL",
            150.0,
            -1500.0,
            None,
            None,
            None,
            "This has to be there",
        ]
    )
    sheet.append(
        [
            None,
            None,
            "Stock purchase",
            datetime(2023, 3, 15, 14, 45),
            "OPEN BUY 5.0 @ 200.00",
            "GOOGL",
            200.0,
            -1000.0,
            None,
            None,
            None,
            None,
        ]
    )

    # Add multiple sells
    sheet.append(
        [
            None,
            None,
            "Stock sale",
            datetime(2023, 4, 10, 11, 0),
            "CLOSE BUY 5.0 @ 160.00",
            "AAPL",
            160.0,
            800.0,
            None,
            None,
            None,
            None,
        ]
    )
    sheet.append(
        [
            None,
            None,
            "Stock sale",
            datetime(2023, 5, 20, 16, 30),
            "CLOSE BUY 2.0 @ 210.00",
            "GOOGL",
            210.0,
            420.0,
            None,
            None,
            None,
            None,
        ]
    )

    # Add multiple dividends
    sheet.append(
        [
            None,
            None,
            "DIVIDENT",
            datetime(2023, 6, 1, 9, 0),
            None,
            None,
            "AAPL",
            10.50,
            None,
            None,
            None,
            None,
        ]
    )
    sheet.append(
        [
            None,
            None,
            "DIVIDENT",
            datetime(2023, 7, 1, 9, 0),
            None,
            None,
            "MSFT",
            8.25,
            None,
            None,
            None,
            None,
        ]
    )

    workbook.save(xlsx_file)

    reader = XtbReader()
    report = reader.read(xlsx_file)

    assert len(report.buys) == 2
    assert report.buys[0].ticker == "AAPL"
    assert report.buys[0].time == "2023-02-01 10:30:00"
    assert report.buys[1].ticker == "GOOGL"
    assert report.buys[1].time == "2023-03-15 14:45:00"

    assert len(report.sells) == 2
    assert report.sells[0].ticker == "AAPL"
    assert report.sells[0].time == "2023-04-10 11:00:00"
    assert report.sells[1].ticker == "GOOGL"
    assert report.sells[1].time == "2023-05-20 16:30:00"

    assert len(report.dividends) == 2
    assert report.dividends[0].ticker == "AAPL"
    assert report.dividends[0].time == "2023-06-01 09:00:00"
    assert report.dividends[0].amount == 10.50
    assert report.dividends[0].currency == Currency.USD
    assert report.dividends[1].ticker == "MSFT"
    assert report.dividends[1].time == "2023-07-01 09:00:00"
    assert report.dividends[1].amount == 8.25
    assert report.dividends[1].currency == Currency.USD


def test_different_deposit_types(tmp_path):
    """Test various deposit types: deposit, Free-funds Interest, Withholding Tax, Stamp Duty."""
    import openpyxl

    xlsx_file = tmp_path / "deposit_types.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("CASH OPERATION HISTORY")

    # Add header rows (must have 12 columns to match data rows)
    sheet.append(
        [None, None, "Currency", None, None, None, "EUR", None, None, None, None, None]
    )
    for _ in range(4):
        sheet.append([None] * 12)
    sheet.append(
        [None, None, "EUR", None, None, "EUR", None, None, None, None, None, None]
    )
    sheet.append(
        [
            None,
            None,
            "Type",
            "Time",
            "Ticker",
            "Amount",
            "Price",
            "Total",
            None,
            None,
            None,
            None,
        ]
    )

    # Regular deposit
    sheet.append(
        [
            None,
            None,
            "deposit",
            datetime(2023, 1, 1, 10, 0),
            None,
            None,
            None,
            1000.1,
            None,
            None,
            None,
            "This has to be there",
        ]
    )
    # Free-funds Interest (adds to deposit)
    sheet.append(
        [
            None,
            None,
            "Free-funds Interest",
            datetime(2023, 2, 1, 10, 0),
            None,
            None,
            None,
            15.50,
            None,
            None,
            None,
            None,
        ]
    )
    # Withholding Tax (negative, but still added)
    sheet.append(
        [
            None,
            None,
            "Withholding Tax",
            datetime(2023, 3, 1, 10, 0),
            None,
            None,
            None,
            -5.25,
            None,
            None,
            None,
            None,
        ]
    )
    # Stamp Duty (negative fee)
    sheet.append(
        [
            None,
            None,
            "Stamp Duty",
            datetime(2023, 4, 1, 10, 0),
            None,
            None,
            None,
            -2.1,
            None,
            None,
            None,
            None,
        ]
    )
    # Free-funds Interest Tax (negative)
    sheet.append(
        [
            None,
            None,
            "Free-funds Interest Tax",
            datetime(2023, 5, 1, 10, 0),
            None,
            None,
            None,
            -1.50,
            None,
            None,
            None,
            None,
        ]
    )

    workbook.save(xlsx_file)

    reader = XtbReader()
    report = reader.read(xlsx_file)

    # All deposit types should be summed: 1000 + 15.50 - 5.25 - 2.00 - 1.50 = 1006.75
    assert report.deposit == 1006.75


def test_transaction_chronological_order(tmp_path):
    """Test that transactions maintain chronological order."""
    import openpyxl

    xlsx_file = tmp_path / "chronological.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("CASH OPERATION HISTORY")

    # Add header rows (must have 12 columns to match data rows)
    sheet.append(
        [None, None, "Currency", None, None, None, "USD", None, None, None, None, None]
    )
    for _ in range(4):
        sheet.append([None] * 12)
    sheet.append(
        [None, None, "EUR", None, None, "EUR", None, None, None, None, None, None]
    )
    sheet.append(
        [
            None,
            None,
            "Type",
            "Time",
            "Ticker",
            "Amount",
            "Price",
            "Total",
            None,
            None,
            None,
            None,
        ]
    )

    # Add transactions in mixed date order
    sheet.append(
        [
            None,
            None,
            "Stock purchase",
            datetime(2023, 3, 15, 10, 0),
            "OPEN BUY 1.0 @ 100.00",
            "STOCK1",
            100.0,
            -100.0,
            None,
            None,
            None,
            "This has to be there",
        ]
    )
    sheet.append(
        [
            None,
            None,
            "Stock purchase",
            datetime(2023, 1, 5, 14, 30),
            "OPEN BUY 2.0 @ 50.00",
            "STOCK2",
            50.0,
            -100.0,
            None,
            None,
            None,
            None,
        ]
    )
    sheet.append(
        [
            None,
            None,
            "Stock purchase",
            datetime(2023, 12, 31, 23, 59),
            "OPEN BUY 3.0 @ 75.00",
            "STOCK3",
            75.0,
            -225.0,
            None,
            None,
            None,
            None,
        ]
    )

    workbook.save(xlsx_file)

    reader = XtbReader()
    report = reader.read(xlsx_file)

    assert len(report.buys) == 3
    # Verify they are in the order they appear in the file
    assert report.buys[0].time == "2023-03-15 10:00:00"
    assert report.buys[1].time == "2023-01-05 14:30:00"
    assert report.buys[2].time == "2023-12-31 23:59:00"


def test_stock_quantity_and_prices(mock_xlsx):
    """Test that stock quantities, share prices, and total prices are correctly parsed."""
    reader = XtbReader()
    report = reader.read(mock_xlsx)

    # Check buy transaction details
    buy = report.buys[0]
    assert buy.ticker == "AAPL"
    assert buy.quantity == 0.1465
    assert buy.share_price == 1023.89  # 150.0 / 0.1465, rounded to 2 decimals
    assert buy.total_price == 150.0

    # Check sell transaction details
    sell = report.sells[0]
    assert sell.ticker == "TSLA"
    assert sell.quantity == 0.0594
    assert sell.share_price == 3367.00  # 200.0 / 0.0594, rounded to 2 decimals
    assert sell.total_price == 200.0


def test_midnight_and_edge_dates(tmp_path):
    """Test transactions at midnight and edge case dates."""
    import openpyxl

    xlsx_file = tmp_path / "edge_dates.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("CASH OPERATION HISTORY")

    # Add header rows (must have 12 columns to match data rows)
    sheet.append(
        [None, None, "Currency", None, None, None, "USD", None, None, None, None, None]
    )
    for _ in range(4):
        sheet.append([None] * 12)
    sheet.append(
        [None, None, None, None, None, None, "EUR", None, None, None, None, None]
    )
    sheet.append(
        [
            None,
            None,
            "Type",
            "Time",
            "Ticker",
            "Amount",
            "Price",
            "Total",
            None,
            None,
            None,
            None,
        ]
    )

    # Transaction at midnight
    sheet.append(
        [
            None,
            None,
            "Stock purchase",
            datetime(2023, 1, 1, 0, 0, 0),
            "OPEN BUY 1.0 @ 100.00",
            "TEST",
            100.0,
            -100.0,
            None,
            None,
            None,
            "This has to be there",
        ]
    )
    # Transaction just before midnight
    sheet.append(
        [
            None,
            None,
            "Stock sale",
            datetime(2023, 12, 31, 23, 59, 59),
            "CLOSE BUY 1.0 @ 100.00",
            "TEST",
            100.0,
            100.0,
            None,
            None,
            None,
            None,
        ]
    )
    # Leap year date
    sheet.append(
        [
            None,
            None,
            "DIVIDENT",
            datetime(2024, 2, 29, 12, 0),
            None,
            None,
            "LEAP",
            5.1,
            None,
            None,
            None,
            None,
        ]
    )

    workbook.save(xlsx_file)

    reader = XtbReader()
    report = reader.read(xlsx_file)

    assert len(report.buys) == 1
    assert report.buys[0].time == "2023-01-01 00:00:00"

    assert len(report.sells) == 1
    assert report.sells[0].time == "2023-12-31 23:59:59"

    assert len(report.dividends) == 1
    assert report.dividends[0].time == "2024-02-29 12:00:00"
    assert report.dividends[0].currency == Currency.EUR
