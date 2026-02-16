import pytest
from datetime import datetime
from convertor.readers.xtb_reader import XtbReader
from convertor.currency import Currency


@pytest.fixture
def mock_xlsx(tmp_path):
    """Creates a temporary XLSX file for testing."""
    import openpyxl

    xlsx_file = tmp_path / "test_xtb.xlsx"
    workbook = openpyxl.Workbook()

    sheet = workbook.create_sheet("CASH OPERATION HISTORY")

    # Add header rows and currency info
    sheet.append([None, None, "Currency", None, None, None, "EUR"])
    sheet.append([None, None, None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, None, None])
    sheet.append([None, None, "EUR", None, None, "EUR", None])

    # Add column headers
    sheet.append([None, None, "Type", "Time", "Ticker", "Amount", "Price", "Total"])

    # Add test data
    sheet.append(
        [
            None,
            None,
            "Stock purchase",
            datetime(2023, 1, 2, 15, 0),
            "OPEN BUY 0.1465 @ 102.36",
            "AAPL",
            150.0,
            -1500.0,
            None,
            None,
            None,
            "This has to be there",
        ]
    )
    sheet.append(
        [
            None,
            None,
            "Stock sale",
            datetime(2023, 1, 3, 16, 0),
            "CLOSE BUY 0.0594/0.104 @ 333.80",
            "TSLA",
            200.0,
            1000.0,
        ]
    )
    sheet.append(
        [None, None, "DIVIDENT", datetime(2023, 1, 4, 9, 0), None, None, "MSFT", 5.50]
    )
    sheet.append(
        [None, None, "deposit", datetime(2023, 1, 1, 10, 0), None, None, None, 1000.1]
    )

    workbook.save(xlsx_file)
    return xlsx_file


def test_xtb_reader_parses_correctly(mock_xlsx):
    reader = XtbReader()

    report = reader.read(mock_xlsx)

    assert report.deposit == 1000.1
    assert len(report.buys) == 1
    assert report.buys[0].ticker == "AAPL"

    assert len(report.sells) == 1
    assert report.sells[0].ticker == "TSLA"

    assert len(report.dividends) == 1
    assert report.dividends[0].ticker == "MSFT"
    assert report.dividends[0].amount == 5.50


def test_currency_extraction(mock_xlsx):
    reader = XtbReader()

    report = reader.read(mock_xlsx)

    assert report.deposit_currency == Currency.EUR


def test_empty_worksheet(tmp_path):
    import openpyxl

    xlsx_file = tmp_path / "empty_xtb.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("CASH OPERATION HISTORY")

    # Add minimal structure
    for i in range(10):
        sheet.append([None] * 8)

    workbook.save(xlsx_file)

    reader = XtbReader()
    report = reader.read(xlsx_file)

    assert report.deposit == 0
    assert len(report.buys) == 0
    assert len(report.sells) == 0
    assert len(report.dividends) == 0
